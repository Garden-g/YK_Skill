#!/usr/bin/env python3
"""Create or update monthly Alibaba inquiry ledger workbooks.

This script is intentionally self-contained because the skill may run in many
different workspaces. It accepts normalized inquiry records as JSON, writes one
Excel workbook per month, deduplicates by customer plus month, and validates
the resulting .xlsx package before reporting success.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, TypeAlias

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as exc:
    OPENPYXL_IMPORT_ERROR: ModuleNotFoundError | None = exc
else:
    OPENPYXL_IMPORT_ERROR = None


JsonDict: TypeAlias = dict[str, Any]
RowDict: TypeAlias = dict[str, str]


VISIBLE_HEADERS: list[str] = [
    "日期",
    "图片",
    "客户",
    "业务",
    "国家",
    "需求",
    "等级",
    "下单情况",
    "公司名称、网站和个人信息",
    "是否添加 WhatsApp，电话沟通等",
    "客户类型",
]

HIDDEN_HEADERS: list[str] = [
    "去重键",
    "会话/询盘来源ID",
    "记录月份",
    "首次记录时间",
    "最后更新时间",
    "筛选原因",
]

ALL_HEADERS: list[str] = VISIBLE_HEADERS + HIDDEN_HEADERS

HEADER_ALIASES: dict[str, str] = {
    "date": "日期",
    "inquiry_date": "日期",
    "image": "图片",
    "image_url": "图片",
    "customer": "客户",
    "buyer": "客户",
    "client": "客户",
    "sales": "业务",
    "owner": "业务",
    "salesperson": "业务",
    "country": "国家",
    "need": "需求",
    "requirement": "需求",
    "message": "需求",
    "level": "等级",
    "grade": "等级",
    "order_status": "下单情况",
    "company_info": "公司名称、网站和个人信息",
    "contact_info": "公司名称、网站和个人信息",
    "whatsapp": "是否添加 WhatsApp，电话沟通等",
    "phone": "是否添加 WhatsApp，电话沟通等",
    "customer_type": "客户类型",
    "type": "客户类型",
}

TEXT_FIELDS_TO_MERGE: list[str] = [
    "日期",
    "图片",
    "业务",
    "国家",
    "需求",
    "下单情况",
    "公司名称、网站和个人信息",
    "是否添加 WhatsApp，电话沟通等",
    "客户类型",
    "会话/询盘来源ID",
    "筛选原因",
]

ADVERTISING_PATTERNS: tuple[str, ...] = (
    "广告",
    "推广服务",
    "营销服务",
    "群发",
    "seo",
    "website design",
    "web design",
    "marketing agency",
    "promotion service",
    "trade show booth",
    "loan",
    "crypto",
    "investment",
)

BANNED_PATTERNS: tuple[str, ...] = (
    "封禁",
    "已封",
    "黑名单",
    "拉黑",
    "风险客户",
    "blocked",
    "banned",
    "blacklist",
)

LOW_QUALITY_PATTERNS: tuple[str, ...] = (
    "低质",
    "无采购意向",
    "非买家",
    "同行推销",
    "无关咨询",
    "招聘",
)

GREETING_ONLY_PATTERNS: tuple[str, ...] = (
    "hi",
    "hello",
    "你好",
    "在吗",
    "您好",
)


@dataclass
class ProcessStats:
    """Store update statistics for a single execution.

    Attributes:
        added: Number of new customer-month rows inserted into workbooks.
        merged: Number of incoming records merged into existing rows.
        filtered: Number of incoming records excluded from the main sheet.
        workbooks: Paths of workbooks that were created or updated.
        filtered_reasons: Human-readable exclusion reasons for audit output.
    """

    added: int = 0
    merged: int = 0
    filtered: int = 0
    workbooks: list[str] | None = None
    filtered_reasons: list[str] | None = None

    def __post_init__(self) -> None:
        """Initialize list attributes after dataclass construction.

        Raises:
            No exceptions are intentionally raised.
        """

        if self.workbooks is None:
            self.workbooks = []
        if self.filtered_reasons is None:
            self.filtered_reasons = []


def parse_args(argv: list[str]) -> argparse.Namespace:
    """Parse command-line arguments.

    Args:
        argv: Raw command-line arguments without the program name.

    Returns:
        Parsed arguments containing input path, output directory, workbook
        prefix, and default year for month/day dates.

    Raises:
        SystemExit: Raised by argparse when required arguments are invalid.
    """

    parser = argparse.ArgumentParser(
        description="Update monthly Alibaba inquiry ledger .xlsx files."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to a JSON file containing a top-level records array.",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory where monthly Excel files should be created or updated.",
    )
    parser.add_argument(
        "--prefix",
        default="询盘记录",
        help="Workbook filename prefix. Default: 询盘记录",
    )
    parser.add_argument(
        "--default-year",
        type=int,
        default=date.today().year,
        help="Year to use when a record date is written as MM.DD or MM-DD.",
    )
    return parser.parse_args(argv)


def load_payload(input_path: Path) -> JsonDict:
    """Load the JSON payload that contains inquiry records.

    Args:
        input_path: Path to the JSON file supplied by the caller.

    Returns:
        Parsed JSON object. A bare list is accepted and converted to
        {"records": list} for convenience.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If the JSON root is not an object or list.
        json.JSONDecodeError: If the file is not valid JSON.
    """

    with input_path.open("r", encoding="utf-8") as file_obj:
        payload = json.load(file_obj)

    if isinstance(payload, list):
        return {"records": payload}
    if not isinstance(payload, dict):
        raise ValueError("Input JSON must be an object or a list of records.")
    return payload


def normalize_text(value: Any) -> str:
    """Convert any cell-like value to clean single-line text.

    Args:
        value: Raw value from JSON or an existing worksheet cell.

    Returns:
        A stripped string with repeated whitespace collapsed. Empty, null, or
        missing values become an empty string.

    Raises:
        No exceptions are intentionally raised.
    """

    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        value = "；".join(normalize_text(item) for item in value if normalize_text(item))
    text = str(value).replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "；", text)
    return text.strip()


def canonical_header(raw_key: str) -> str:
    """Map aliases such as English keys into the Chinese ledger headers.

    Args:
        raw_key: Key found in the incoming JSON record.

    Returns:
        Canonical header name when an alias is known; otherwise the original
        key is returned unchanged.

    Raises:
        No exceptions are intentionally raised.
    """

    key = normalize_text(raw_key)
    return HEADER_ALIASES.get(key, HEADER_ALIASES.get(key.lower(), key))


def normalize_record(raw_record: JsonDict, default_year: int) -> RowDict:
    """Normalize one raw inquiry object into the workbook row shape.

    Args:
        raw_record: Incoming record from the JSON payload.
        default_year: Year used when the record date omits a year.

    Returns:
        Dictionary containing every visible and hidden workbook header.

    Raises:
        ValueError: If the date cannot be parsed into a usable month.
    """

    row: RowDict = {header: "" for header in ALL_HEADERS}

    for raw_key, raw_value in raw_record.items():
        header = canonical_header(raw_key)
        if header in row:
            row[header] = normalize_text(raw_value)

    source_id = normalize_text(
        raw_record.get("source_id")
        or raw_record.get("inquiry_id")
        or raw_record.get("conversation_id")
        or raw_record.get("会话ID")
        or raw_record.get("询盘ID")
    )
    row["会话/询盘来源ID"] = merge_unique_text(row["会话/询盘来源ID"], source_id)

    parsed_date = parse_record_date(row["日期"], default_year)
    row["日期"] = parsed_date.isoformat()
    row["记录月份"] = parsed_date.strftime("%Y-%m")

    customer = normalize_text(row["客户"])
    if not customer and source_id:
        customer = f"未知客户-{source_id}"
        row["客户"] = customer
    if not customer:
        raise ValueError("Record is missing 客户/customer and source_id.")

    row["去重键"] = build_dedupe_key(customer, row["记录月份"])
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row["首次记录时间"] = normalize_text(raw_record.get("首次记录时间")) or now_text
    row["最后更新时间"] = now_text

    filter_reason = detect_filter_reason(raw_record, row)
    row["筛选原因"] = filter_reason
    return row


def parse_record_date(raw_date: str, default_year: int) -> date:
    """Parse common inquiry date formats into a Python date.

    Args:
        raw_date: Date text from the record, such as 2026-04-01, 04.01, or
            4/1/2026.
        default_year: Year used for month/day-only formats.

    Returns:
        Parsed date object.

    Raises:
        ValueError: If the date is empty or does not match supported formats.
    """

    text = normalize_text(raw_date)
    if not text:
        raise ValueError("Record date is required.")

    candidates = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%m-%d-%Y",
        "%m/%d/%Y",
        "%m.%d.%Y",
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    month_day_match = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})", text)
    if month_day_match:
        month = int(month_day_match.group(1))
        day = int(month_day_match.group(2))
        return date(default_year, month, day)

    raise ValueError(f"Unsupported date format: {raw_date}")


def build_dedupe_key(customer: str, month: str) -> str:
    """Build the stable customer-month key used for deduplication.

    Args:
        customer: Customer or buyer name.
        month: Month string in YYYY-MM format.

    Returns:
        Lowercase normalized key in the form "month::customer".

    Raises:
        No exceptions are intentionally raised.
    """

    normalized_customer = re.sub(r"\s+", " ", customer).strip().lower()
    return f"{month}::{normalized_customer}"


def detect_filter_reason(raw_record: JsonDict, row: RowDict) -> str:
    """Decide whether an incoming inquiry should be excluded from the main sheet.

    Args:
        raw_record: Original incoming record, including optional explicit flags.
        row: Normalized row data.

    Returns:
        Empty string when the record should be kept; otherwise a short Chinese
        reason such as "广告推广" or "已封禁/黑名单".

    Raises:
        No exceptions are intentionally raised.
    """

    explicit_banned = bool(raw_record.get("is_banned") or raw_record.get("banned"))
    explicit_ad = bool(raw_record.get("is_advertising") or raw_record.get("advertising"))
    explicit_low = bool(raw_record.get("is_low_quality") or raw_record.get("low_quality"))

    combined_text = " ".join(row.values()).lower()
    if explicit_banned or contains_any(combined_text, BANNED_PATTERNS):
        return "已封禁/黑名单"
    if explicit_ad or contains_any(combined_text, ADVERTISING_PATTERNS):
        return "广告推广/群发营销"
    if explicit_low or contains_any(combined_text, LOW_QUALITY_PATTERNS):
        return "低质/非采购意向"
    if is_greeting_only(row["需求"]):
        return "仅寒暄且无采购需求"
    return ""


def contains_any(text: str, patterns: Iterable[str]) -> bool:
    """Check whether text contains any configured filter pattern.

    Args:
        text: Lowercase text to inspect.
        patterns: Words or phrases that indicate a filter reason.

    Returns:
        True when any pattern appears in the text.

    Raises:
        No exceptions are intentionally raised.
    """

    return any(pattern.lower() in text for pattern in patterns)


def is_greeting_only(requirement: str) -> bool:
    """Detect messages that are only greetings with no purchase signal.

    Args:
        requirement: Normalized customer requirement text.

    Returns:
        True when the message is only a short greeting; false when the message
        contains any likely product or purchase detail.

    Raises:
        No exceptions are intentionally raised.
    """

    text = re.sub(r"[!！?.。,\s]+", "", requirement.lower())
    if not text:
        return True
    return text in {re.sub(r"\s+", "", item.lower()) for item in GREETING_ONLY_PATTERNS}


def merge_unique_text(existing: str, incoming: str) -> str:
    """Merge semicolon-separated text while preserving unique values.

    Args:
        existing: Current cell text.
        incoming: New cell text.

    Returns:
        Combined text with duplicate segments removed and original order kept.

    Raises:
        No exceptions are intentionally raised.
    """

    parts: list[str] = []
    seen: set[str] = set()
    for source in (existing, incoming):
        for piece in re.split(r"[；;]\s*", normalize_text(source)):
            normalized_piece = normalize_text(piece)
            key = normalized_piece.lower()
            if normalized_piece and key not in seen:
                seen.add(key)
                parts.append(normalized_piece)
    return "；".join(parts)


def choose_better_level(existing: str, incoming: str) -> str:
    """Keep the highest known customer level when merging duplicate rows.

    Args:
        existing: Current level text from the workbook.
        incoming: New level text from the incoming record.

    Returns:
        The higher level when both values look like L1/L2/L3; otherwise unique
        text values are merged.

    Raises:
        No exceptions are intentionally raised.
    """

    existing_clean = normalize_text(existing).upper()
    incoming_clean = normalize_text(incoming).upper()
    level_rank = {"L0": 0, "L1": 1, "L2": 2, "L3": 3, "L4": 4}
    if existing_clean in level_rank and incoming_clean in level_rank:
        return existing_clean if level_rank[existing_clean] >= level_rank[incoming_clean] else incoming_clean
    return merge_unique_text(existing, incoming)


def workbook_path(output_dir: Path, prefix: str, month: str) -> Path:
    """Build the workbook path for one record month.

    Args:
        output_dir: Directory where monthly workbooks live.
        prefix: Filename prefix, normally "询盘记录".
        month: Month in YYYY-MM format.

    Returns:
        Full path like output_dir/询盘记录_2026-04.xlsx.

    Raises:
        No exceptions are intentionally raised.
    """

    return output_dir / f"{prefix}_{month}.xlsx"


def load_or_create_workbook(path: Path) -> Workbook:
    """Open an existing workbook or create a new inquiry ledger workbook.

    Args:
        path: Monthly workbook path.

    Returns:
        An openpyxl Workbook with a configured active sheet.

    Raises:
        openpyxl exceptions: Propagated when an existing workbook is corrupt.
    """

    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook.active
        ensure_headers(worksheet)
        return workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "询盘记录"
    ensure_headers(worksheet)
    return workbook


def ensure_headers(worksheet: Worksheet) -> None:
    """Ensure the worksheet has the required visible and hidden headers.

    Args:
        worksheet: Sheet to initialize or repair.

    Returns:
        None. The worksheet is modified in place.

    Raises:
        No exceptions are intentionally raised.
    """

    for col_index, header in enumerate(ALL_HEADERS, start=1):
        worksheet.cell(row=1, column=col_index, value=header)
    apply_sheet_style(worksheet)


def apply_sheet_style(worksheet: Worksheet) -> None:
    """Apply the standard ledger layout and visual formatting.

    Args:
        worksheet: Sheet whose columns, header, freeze panes, and hidden helper
            columns should be styled.

    Returns:
        None. The worksheet is modified in place.

    Raises:
        No exceptions are intentionally raised.
    """

    header_fill = PatternFill("solid", fgColor="FFF200")
    hidden_fill = PatternFill("solid", fgColor="D9EAD3")
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    widths = {
        "日期": 14,
        "图片": 20,
        "客户": 24,
        "业务": 16,
        "国家": 14,
        "需求": 60,
        "等级": 10,
        "下单情况": 18,
        "公司名称、网站和个人信息": 42,
        "是否添加 WhatsApp，电话沟通等": 28,
        "客户类型": 22,
    }

    for col_index, header in enumerate(ALL_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.font = Font(name="Arial", bold=True, color="000000")
        cell.fill = header_fill if header in VISIBLE_HEADERS else hidden_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        column_letter = get_column_letter(col_index)
        worksheet.column_dimensions[column_letter].width = widths.get(header, 22)
        if header in HIDDEN_HEADERS:
            worksheet.column_dimensions[column_letter].hidden = True

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(ALL_HEADERS))}{max(worksheet.max_row, 1)}"


def read_existing_rows(worksheet: Worksheet) -> dict[str, RowDict]:
    """Read existing worksheet rows into a dictionary keyed by dedupe key.

    Args:
        worksheet: Existing monthly inquiry worksheet.

    Returns:
        Mapping from hidden dedupe key to row dictionaries.

    Raises:
        No exceptions are intentionally raised.
    """

    rows_by_key: dict[str, RowDict] = {}
    for row_index in range(2, worksheet.max_row + 1):
        row = {
            header: normalize_text(worksheet.cell(row=row_index, column=col_index).value)
            for col_index, header in enumerate(ALL_HEADERS, start=1)
        }
        key = row.get("去重键") or build_dedupe_key(row.get("客户", ""), row.get("记录月份", ""))
        if key.strip():
            row["去重键"] = key
            rows_by_key[key] = row
    return rows_by_key


def merge_record(existing: RowDict, incoming: RowDict) -> RowDict:
    """Merge an incoming duplicate customer-month record into an existing row.

    Args:
        existing: Current workbook row for the customer-month key.
        incoming: New normalized row for the same key.

    Returns:
        Merged row dictionary. The original existing dictionary is not mutated.

    Raises:
        No exceptions are intentionally raised.
    """

    merged = dict(existing)
    for field in TEXT_FIELDS_TO_MERGE:
        merged[field] = merge_unique_text(existing.get(field, ""), incoming.get(field, ""))

    merged["等级"] = choose_better_level(existing.get("等级", ""), incoming.get("等级", ""))
    merged["客户"] = existing.get("客户") or incoming.get("客户", "")
    merged["去重键"] = existing.get("去重键") or incoming.get("去重键", "")
    merged["记录月份"] = existing.get("记录月份") or incoming.get("记录月份", "")
    merged["首次记录时间"] = existing.get("首次记录时间") or incoming.get("首次记录时间", "")
    merged["最后更新时间"] = incoming.get("最后更新时间") or existing.get("最后更新时间", "")
    return merged


def first_date_for_sort(row: RowDict) -> date:
    """Find the first valid date inside a merged row for chronological sorting.

    Args:
        row: Workbook row dictionary.

    Returns:
        Parsed date used for sorting; date.max is returned when no date parses.

    Raises:
        No exceptions are intentionally raised.
    """

    for piece in re.split(r"[；;]\s*", row.get("日期", "")):
        try:
            return parse_record_date(piece, date.today().year)
        except ValueError:
            continue
    return date.max


def rewrite_rows(worksheet: Worksheet, rows: Iterable[RowDict]) -> None:
    """Rewrite worksheet data rows in chronological order.

    Args:
        worksheet: Target worksheet.
        rows: Row dictionaries to write below the header.

    Returns:
        None. Existing data rows are removed and replaced.

    Raises:
        No exceptions are intentionally raised.
    """

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    sorted_rows = sorted(rows, key=first_date_for_sort)
    for row_offset, row in enumerate(sorted_rows, start=2):
        for col_index, header in enumerate(ALL_HEADERS, start=1):
            cell = worksheet.cell(row=row_offset, column=col_index, value=row.get(header, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", color="000000")
        style_data_row(worksheet, row_offset, row)

    apply_sheet_style(worksheet)


def style_data_row(worksheet: Worksheet, row_index: int, row: RowDict) -> None:
    """Apply row-level styling for merged or pending-confirmation records.

    Args:
        worksheet: Sheet containing the row.
        row_index: One-based row number to style.
        row: Row dictionary, used to decide highlight color.

    Returns:
        None. The row cells are styled in place.

    Raises:
        No exceptions are intentionally raised.
    """

    thin_side = Side(style="thin", color="E6E6E6")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    date_text = row.get("日期", "")
    needs_fill = "待确认" in " ".join(row.values())
    is_merged = "；" in date_text or "；" in row.get("需求", "")
    fill = None
    if needs_fill:
        fill = PatternFill("solid", fgColor="FFF2CC")
    elif is_merged:
        fill = PatternFill("solid", fgColor="FCE4D6")

    for col_index in range(1, len(ALL_HEADERS) + 1):
        cell = worksheet.cell(row=row_index, column=col_index)
        cell.border = border
        if fill is not None:
            cell.fill = copy(fill)


def update_workbooks(records: list[RowDict], output_dir: Path, prefix: str) -> ProcessStats:
    """Apply normalized records to their monthly workbooks.

    Args:
        records: Normalized and already-filtered records.
        output_dir: Directory where workbooks should be stored.
        prefix: Workbook filename prefix.

    Returns:
        ProcessStats with counts and touched workbook paths.

    Raises:
        OSError: If output files cannot be written.
        openpyxl exceptions: If existing workbook files cannot be opened.
    """

    stats = ProcessStats()
    records_by_month: dict[str, list[RowDict]] = {}
    for record in records:
        records_by_month.setdefault(record["记录月份"], []).append(record)

    output_dir.mkdir(parents=True, exist_ok=True)

    for month, month_records in sorted(records_by_month.items()):
        path = workbook_path(output_dir, prefix, month)
        workbook = load_or_create_workbook(path)
        worksheet = workbook.active
        rows_by_key = read_existing_rows(worksheet)

        for record in month_records:
            key = record["去重键"]
            if key in rows_by_key:
                rows_by_key[key] = merge_record(rows_by_key[key], record)
                stats.merged += 1
            else:
                rows_by_key[key] = record
                stats.added += 1

        rewrite_rows(worksheet, rows_by_key.values())
        workbook.save(path)
        validate_xlsx_package(path)
        stats.workbooks.append(str(path))

    return stats


def validate_xlsx_package(path: Path) -> None:
    """Validate that an .xlsx workbook opens and has no table/drawing residue.

    Args:
        path: Workbook path to validate.

    Returns:
        None when validation succeeds.

    Raises:
        ValueError: If package validation finds table/drawing residue.
        zipfile.BadZipFile: If the workbook is not a valid zip package.
        openpyxl exceptions: If the workbook cannot be loaded.
    """

    with zipfile.ZipFile(path, "r") as archive:
        bad_file = archive.testzip()
        if bad_file:
            raise ValueError(f"Invalid zipped file inside workbook: {bad_file}")
        names = archive.namelist()
        residue = [
            name
            for name in names
            if name.startswith("xl/tables/")
            or name.startswith("xl/drawings/")
            or ("table" in name.lower() and name.endswith(".rels"))
            or ("drawing" in name.lower() and name.endswith(".rels"))
        ]
        if residue:
            raise ValueError(f"Unexpected table/drawing residue found: {residue}")

    workbook = load_workbook(path)
    workbook.close()


def process_payload(payload: JsonDict, output_dir: Path, prefix: str, default_year: int) -> ProcessStats:
    """Normalize, filter, and write records from a parsed JSON payload.

    Args:
        payload: Parsed input JSON containing a records array.
        output_dir: Directory where monthly workbooks should be written.
        prefix: Workbook filename prefix.
        default_year: Year used for MM.DD-style dates.

    Returns:
        ProcessStats with update counts and touched workbook paths.

    Raises:
        ValueError: If the records field is missing or malformed, or if no valid
            records remain after filtering.
    """

    raw_records = payload.get("records")
    if not isinstance(raw_records, list):
        raise ValueError("Input JSON must contain a 'records' array.")

    stats = ProcessStats()
    kept_records: list[RowDict] = []
    for index, raw_record in enumerate(raw_records, start=1):
        if not isinstance(raw_record, dict):
            stats.filtered += 1
            stats.filtered_reasons.append(f"第 {index} 条：记录不是对象")
            continue
        try:
            row = normalize_record(raw_record, default_year)
        except ValueError as exc:
            stats.filtered += 1
            stats.filtered_reasons.append(f"第 {index} 条：{exc}")
            continue

        if row["筛选原因"]:
            stats.filtered += 1
            stats.filtered_reasons.append(f"{row['客户']}：{row['筛选原因']}")
            continue
        kept_records.append(row)

    if kept_records:
        write_stats = update_workbooks(kept_records, output_dir, prefix)
        stats.added += write_stats.added
        stats.merged += write_stats.merged
        stats.workbooks.extend(write_stats.workbooks)

    if not kept_records and stats.filtered == 0:
        raise ValueError("No records were provided.")

    return stats


def main(argv: list[str]) -> int:
    """Run the monthly inquiry ledger updater.

    Args:
        argv: Command-line arguments without the program name.

    Returns:
        Process exit code. Zero means success; non-zero means the workbook was
        not safely produced.

    Raises:
        No exceptions are intentionally raised; user-facing failures are printed
        to stderr and converted to exit code 1.
    """

    try:
        if OPENPYXL_IMPORT_ERROR is not None:
            raise RuntimeError(
                "Missing dependency: openpyxl. Run this script with a Python "
                "environment that includes openpyxl before generating .xlsx files."
            ) from OPENPYXL_IMPORT_ERROR
        args = parse_args(argv)
        payload = load_payload(Path(args.input))
        stats = process_payload(payload, Path(args.output_dir), args.prefix, args.default_year)
    except Exception as exc:  # noqa: BLE001 - CLI should convert all failures to a clear message.
        print(json.dumps({"status": "failed", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1

    print(
        json.dumps(
            {
                "status": "success",
                "added": stats.added,
                "merged": stats.merged,
                "filtered": stats.filtered,
                "workbooks": stats.workbooks,
                "filtered_reasons": stats.filtered_reasons,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
